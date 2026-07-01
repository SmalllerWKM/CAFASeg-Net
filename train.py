import os
import argparse
import logging
import math
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterator, List, Optional, Tuple

import torch
import torch.nn as nn
from torch import Tensor
from torch.utils.data import DataLoader
from torch.utils.tensorboard import SummaryWriter
from tqdm import tqdm


try:
    from tabulate import tabulate
    _HAS_TABULATE = True
except ImportError:
    _HAS_TABULATE = False
    def tabulate(rows, headers=(), tablefmt="simple", floatfmt=".4f", **kw):
        all_rows = [list(headers)] + [list(r) for r in rows] if headers else [list(r) for r in rows]
        col_w = [max(len(str(cell)) for cell in col) for col in zip(*all_rows)]
        sep = "-+-".join("-" * w for w in col_w)
        lines = []
        for i, row in enumerate(all_rows):
            lines.append(" | ".join(str(v).ljust(w) for v, w in zip(row, col_w)))
            if i == 0 and headers:
                lines.append(sep)
        return "\n".join(lines)


try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


try:
    from torch.amp import GradScaler, autocast as amp_autocast
    _AMP_DEVICE = "cuda"
except ImportError:
    from torch.cuda.amp import GradScaler, autocast as amp_autocast
    _AMP_DEVICE = None

def autocast(enabled: bool = True):
    if _AMP_DEVICE:
        return amp_autocast(device_type=_AMP_DEVICE, enabled=enabled)
    return amp_autocast(enabled=enabled)

from config  import Config, cfg, require_runtime_assets
from dataset import build_labeled_loaders
from models  import CAFASegNet
from losses  import CAFASegNetLoss
from inference import keep_largest_component_tensor
from metrics import SegMetrics


RESET  = "\033[0m"
BOLD   = "\033[1m"
CYAN   = "\033[36m"
GREEN  = "\033[32m"
YELLOW = "\033[33m"
RED    = "\033[31m"
BLUE   = "\033[34m"
GREY   = "\033[90m"


def setup_logging(log_dir: str) -> logging.Logger:
    Path(log_dir).mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger("CAFASegNet")
    logger.setLevel(logging.INFO)
    if logger.handlers:
        return logger
    fh  = logging.FileHandler(os.path.join(log_dir, "train.log"), encoding="utf-8")
    ch  = logging.StreamHandler()
    fmt = logging.Formatter("[%(asctime)s] %(levelname)s — %(message)s", "%Y-%m-%d %H:%M:%S")
    fh.setFormatter(fmt)
    ch.setFormatter(fmt)
    logger.addHandler(fh)
    logger.addHandler(ch)
    return logger


def set_seed(seed: int) -> None:
    import random, numpy as np
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark     = False

def cosine_anneal_tau(epoch: int, tau_init: float, tau_min: float, anneal_epochs: int) -> float:
    if epoch >= anneal_epochs:
        return tau_min
    ratio = epoch / anneal_epochs
    return tau_min + 0.5 * (tau_init - tau_min) * (1.0 + math.cos(math.pi * ratio))

def make_infinite(loader: DataLoader) -> Iterator:
    while True:
        yield from loader

def get_lr(optimizer: torch.optim.Optimizer) -> float:
    return optimizer.param_groups[0]["lr"]

def is_valid_loss(loss: Tensor) -> bool:
    return bool(torch.isfinite(loss).all())

def build_scheduler(optimizer, cfg_train) -> torch.optim.lr_scheduler.SequentialLR:
    warmup_epochs = cfg_train.warmup_epochs
    total_epochs  = cfg_train.num_epochs
    warmup_scheduler = torch.optim.lr_scheduler.LinearLR(
        optimizer, start_factor=1e-3, end_factor=1.0, total_iters=warmup_epochs,
    )
    cosine_scheduler = torch.optim.lr_scheduler.CosineAnnealingLR(
        optimizer, T_max=total_epochs - warmup_epochs, eta_min=cfg_train.lr_min,
    )
    return torch.optim.lr_scheduler.SequentialLR(
        optimizer, schedulers=[warmup_scheduler, cosine_scheduler], milestones=[warmup_epochs],
    )


def save_checkpoint(state: dict, path: str, is_best: bool = False, best_path: str = "") -> None:
    torch.save(state, path)
    if is_best and best_path:
        import shutil
        shutil.copyfile(path, best_path)

def load_checkpoint(model, optimizer, scheduler, scaler, path: str, device: torch.device, logger: logging.Logger) -> Tuple[int, float]:
    logger.info(f"Loading checkpoint: {path}")
    ckpt = torch.load(path, map_location=device)
    model.load_state_dict(ckpt["model"])
    optimizer.load_state_dict(ckpt["optimizer"])
    scheduler.load_state_dict(ckpt["scheduler"])
    if "scaler" in ckpt:
        scaler.load_state_dict(ckpt["scaler"])

    start_epoch = ckpt.get("epoch", 0) + 1
    best_dice   = ckpt.get("best_dice", 0.0)
    logger.info(f"Resumed from epoch {start_epoch} (best Dice: {best_dice:.4f})")
    return start_epoch, best_dice


def _fmt(v, fmt=".4f") -> str:
    if isinstance(v, float):
        if math.isnan(v): return "—"
        if math.isinf(v): return "∞"
        return f"{v:{fmt}}"
    return str(v)

def print_startup_table(cfg: Config, device: torch.device) -> None:
    print(f"\n{BOLD}{CYAN}{'━'*64}{RESET}")
    print(f"{BOLD}{CYAN}   CAFASeg-Net (DSBA-Bone) Training Startup{RESET}")
    print(f"{BOLD}{CYAN}{'━'*64}{RESET}")
    _bb_lr_scale  = getattr(cfg.train, "backbone_lr_scale", 1.0)

    rows = [
        ["Device",    str(device)],
        ["AMP", str(cfg.train.amp)],
        ["Epochs",    str(cfg.train.num_epochs)],
        ["Batch Size",   str(cfg.train.batch_size)],
        ["Backbone",     "DSBA-Bone (ResNet34 + boundary stream)"],
        ["Backbone LR scale", f"x{_bb_lr_scale}" if _bb_lr_scale != 1.0 else "—"],
        ["Initial LR",   f"{cfg.train.lr:.2e}"],
        ["Warmup epochs",   str(cfg.train.warmup_epochs)],
        ["Decoder channels",  str(cfg.model.decoder_channels)],
        ["lambda_bce",   str(cfg.train.lambda_bce)],
        ["lambda_tversky", str(cfg.train.lambda_tversky)],
    ]
    print(tabulate(rows, headers=["Setting", "Value"], tablefmt="rounded_outline"))
    print()

def print_train_table(epoch: int, total: int, stage: str, avg: Dict[str, float], tau: float, lr: float, elapsed: float) -> None:
    title = (
        f"\n{BOLD}  Epoch {epoch:03d}/{total}{RESET}"
        f"  [{BOLD}{GREEN}{stage}{RESET}]"
        f"  τ={GREY}{tau:.3f}{RESET}"
        f"  LR={GREY}{lr:.2e}{RESET}"
        f"  {GREY}{elapsed:.0f}s{RESET}"
    )
    print(title)
    rows = [
        ["Total Loss",      _fmt(avg["total"]),          "BW-Dice",    _fmt(avg.get("bw_dice", float("nan")))],
        ["Tversky Loss", _fmt(avg["tversky_loss"]),    "Deep supervision",   _fmt(avg.get("deep_sup", 0.0))],
        ["CC-Dice Loss", _fmt(avg.get("cc_dice_loss", 0.0)), "NaN steps", str(int(avg["nan_steps"]))],
        ["Focal Loss",   _fmt(avg["bce_loss"]),        "", ""],
    ]
    print(tabulate(rows, headers=["Training metrics", "Value", "Auxiliary metrics", "Value"], tablefmt="rounded_outline"))

def print_val_table(epoch: int, result: Dict) -> None:
    dice = result.get("dice", 0.0)
    if   dice >= 0.75: dice_color = BOLD + GREEN
    elif dice >= 0.60: dice_color = BOLD + YELLOW
    else:              dice_color = BOLD + RED

    print(f"\n{BOLD}  ▶ Epoch {epoch:03d} Validation results{RESET}")
    valid_str = f"{result['n_valid']}/{result['n_total']}"
    cr      = result.get("comp_recall", float("nan"))
    cr_str = f"{cr:.4f}" if not math.isnan(cr) else "N/A"
    cr_detail = f"{result.get('comp_hit',0)}/{result.get('comp_total',0)}"

    rows = [
        ["Dice ↑",       f"{dice_color}{result['dice']:.4f}{RESET}", _fmt(result["dice_std"]), valid_str, str(result["n_inf"])],
        ["IoU ↑",        _fmt(result["iou"]),        _fmt(result["iou_std"]), "", ""],
        ["HD95 ↓ (mm)",  _fmt(result["hd95"], ".2f"), _fmt(result["hd95_std"], ".2f"), "", ""],
        ["Component Recall ↑",      cr_str, "", cr_detail, ""],
        ["Val Loss",     _fmt(result["val_loss"]),   "", f"skip={result['n_skip']}", ""],
    ]
    print(tabulate(rows, headers=["Metric", "Mean", "Std", "Valid/Total", "Inf frames"], tablefmt="rounded_outline"))

def print_best_banner(epoch: int, dice: float) -> None:
    bar = f"{BOLD}{GREEN}{'★' * 22}{RESET}"
    print(f"\n  {bar}")
    print(f"  {BOLD}{GREEN}  ★  New best Dice: {dice:.4f}  —  Epoch {epoch:03d}  ★{RESET}")
    print(f"  {bar}")


_EXCEL_COLS: List[Tuple[str, int]] = [
    ("Epoch", 8), ("Stage", 8), ("Train_Loss", 12), ("Dice_Loss", 12),
    ("Tversky_Loss", 14), ("CC_Dice_Loss", 14), ("Deep_Sup_Loss", 14),
    ("Focal_Loss", 12), ("Val_Loss", 12), ("Val_Dice", 12), ("Val_Dice_Std", 14),
    ("Val_IoU", 10), ("Val_IoU_Std", 12), ("Val_HD95_mm", 14), ("Val_HD95_Std", 14),
    ("Valid_Frames", 14), ("Total_Frames", 14), ("Skip_Frames", 12),
    ("Inf_Frames", 12), ("LR", 14), ("Tau", 10), ("Elapsed_s", 12), ("Timestamp", 22),
]
_FMT4 = {"Train_Loss", "Dice_Loss", "Tversky_Loss", "Focal_Loss", "CC_Dice_Loss", "Deep_Sup_Loss", "Val_Loss", "Val_Dice", "Val_Dice_Std", "Val_IoU", "Val_IoU_Std", "Tau"}
_FMT2 = {"Val_HD95_mm", "Val_HD95_Std", "Elapsed_s"}
_FMTE = {"LR"}

class ExcelLogger:
    def __init__(self, log_dir: str, filename: str = "val_results.xlsx") -> None:
        self.path      = os.path.join(log_dir, filename)
        self.best_dice = 0.0
        self.best_row  = -1
        self._wb       = None
        self._ws       = None

        if not _HAS_OPENPYXL: return

        if os.path.exists(self.path):
            self._wb = openpyxl.load_workbook(self.path)
            self._ws = self._wb.active
            for row in self._ws.iter_rows(min_row=2, values_only=True):
                val = row[9]
                if val is not None:
                    try:
                        v = float(val)
                        if v > self.best_dice: self.best_dice = v
                    except (TypeError, ValueError): pass
        else:
            self._wb = openpyxl.Workbook()
            self._ws = self._wb.active
            self._ws.title = "Validation results"
            self._write_header()
            self._wb.save(self.path)

    @staticmethod
    def _border():
        side = Side(style="thin", color="BDD7EE")
        return Border(left=side, right=side, top=side, bottom=side)

    def _write_header(self) -> None:
        ws = self._ws
        hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        hdr_fill  = PatternFill("solid", fgColor="1F4E79")
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for ci, (col_name, col_w) in enumerate(_EXCEL_COLS, start=1):
            cell = ws.cell(row=1, column=ci, value=col_name)
            cell.font, cell.fill, cell.alignment, cell.border = hdr_font, hdr_fill, hdr_align, self._border()
            ws.column_dimensions[get_column_letter(ci)].width = col_w
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"

    def _style_row(self, row_num: int, is_best: bool) -> None:
        ws = self._ws
        bg_color = "E2EFDA" if is_best else ("F2F7FB" if row_num % 2 == 0 else "FFFFFF")
        fill     = PatternFill("solid", fgColor=bg_color)
        border   = self._border()
        font     = Font(name="Arial", bold=is_best, size=9, color="1F4E79" if is_best else "000000")

        for ci, (col_name, _) in enumerate(_EXCEL_COLS, start=1):
            cell = ws.cell(row=row_num, column=ci)
            cell.fill, cell.border, cell.font = fill, border, font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_name in _FMT4: cell.number_format = "0.0000"
            elif col_name in _FMT2: cell.number_format = "0.00"
            elif col_name in _FMTE: cell.number_format = "0.00E+00"

    def _clear_best_row(self) -> None:
        if self.best_row >= 2: self._style_row(self.best_row, is_best=False)

    def append(self, epoch: int, stage: str, train_avg: Dict[str, float], val_result: Dict, lr: float, tau: float, elapsed: float) -> None:
        if not _HAS_OPENPYXL or self._wb is None: return
        current_dice = float(val_result.get("dice", 0.0))
        is_new_best  = current_dice > self.best_dice
        if is_new_best:
            self._clear_best_row()
            self.best_dice = current_dice
            self.best_row  = self._ws.max_row + 1

        data = [
            epoch, stage,
            train_avg.get("total", float("nan")), train_avg.get("dice_loss", float("nan")),
            train_avg.get("tversky_loss", float("nan")), train_avg.get("cc_dice_loss", float("nan")),
            train_avg.get("deep_sup", float("nan")), train_avg.get("bce_loss", float("nan")),
            val_result.get("val_loss", float("nan")), val_result.get("dice", float("nan")),
            val_result.get("dice_std", float("nan")), val_result.get("iou", float("nan")),
            val_result.get("iou_std", float("nan")), val_result.get("hd95", float("nan")),
            val_result.get("hd95_std", float("nan")), val_result.get("n_valid", 0),
            val_result.get("n_total", 0), val_result.get("n_skip", 0), val_result.get("n_inf", 0),
            lr, tau, elapsed, datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ]
        row_num = self._ws.max_row + 1
        for ci, value in enumerate(data, start=1):
            self._ws.cell(row=row_num, column=ci, value=value)
        self._style_row(row_num, is_new_best)
        self._wb.save(self.path)

    def close(self) -> None:
        if _HAS_OPENPYXL and self._wb is not None: self._wb.save(self.path)


def train_one_epoch(
    model: CAFASegNet, loader: DataLoader, criterion: CAFASegNetLoss,
    optimizer: torch.optim.Optimizer, scaler: GradScaler, device: torch.device,
    use_amp: bool, grad_clip: float, writer: SummaryWriter, global_step: int,
    epoch: int, ema: "ModelEMA | None" = None,
) -> Tuple[int, Dict[str, float]]:
    model.train()
    accum: Dict[str, float] = dict.fromkeys(["dice_loss", "tversky_loss", "bce_loss", "bw_dice", "cc_dice_loss", "deep_sup", "total"], 0.0)
    n_valid = 0
    n_nan   = 0

    pbar = tqdm(loader, desc=f"  {CYAN}Train{RESET} E{epoch:03d}", ncols=130, leave=False, dynamic_ncols=False,
                bar_format=("  {desc}: {percentage:3.0f}%|{bar:30}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}]  {postfix}"))

    for batch in pbar:
        images  = batch["image"].to(device, non_blocking=True)
        targets = batch["mask"].to(device, non_blocking=True)

        with autocast(enabled=use_amp):
            out      = model(images, return_features=False)
            sup_dict = criterion.supervised_loss(out, targets, model)
            loss     = sup_dict["total"]

        if not is_valid_loss(loss):
            n_nan += 1
            optimizer.zero_grad()
            global_step += 1
            pbar.set_postfix({"⚠ NaN": n_nan}, refresh=False)
            continue

        optimizer.zero_grad()
        scaler.scale(loss).backward()
        scaler.unscale_(optimizer)
        nn.utils.clip_grad_norm_(model.parameters(), grad_clip)
        scaler.step(optimizer)
        scaler.update()
        if ema: ema.update(model)

        for k in accum.keys():
            accum[k] += sup_dict.get(k, torch.tensor(0.0)).item()

        n_valid    += 1
        global_step += 1

        pbar.set_postfix({
            "loss":    f"{sup_dict['total'].item():.4f}",
            "dice":    f"{sup_dict['dice_loss'].item():.4f}",
            "tversky": f"{sup_dict['tversky_loss'].item():.4f}",
        }, refresh=False)

        if global_step % 20 == 0:
            writer.add_scalar("Step/Loss_total", sup_dict["total"].item(), global_step)
            writer.add_scalar("Step/Loss_dice", sup_dict["dice_loss"].item(), global_step)
            writer.add_scalar("Step/Loss_focal", sup_dict["bce_loss"].item(), global_step)
            writer.add_scalar("Step/LR", get_lr(optimizer), global_step)

    pbar.close()
    denom = max(n_valid, 1)
    avg   = {k: v / denom for k, v in accum.items()}
    avg["nan_steps"] = float(n_nan)
    return global_step, avg


@torch.no_grad()
def validate(model: CAFASegNet, loader: DataLoader, criterion: CAFASegNetLoss, device: torch.device, use_amp: bool, epoch: int, ema: "ModelEMA | None" = None) -> Dict:
    model.eval()
    metrics  = SegMetrics(compute_hd=True)
    val_loss = 0.0
    n_steps  = 0

    pbar = tqdm(loader, desc=f"  {BLUE}Val{RESET} E{epoch:03d}", ncols=130, leave=False, dynamic_ncols=False,
                bar_format=("  {desc}: {percentage:3.0f}%|{bar:30}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}]  {postfix}"))

    for batch in pbar:
        images  = batch["image"].to(device)
        targets = batch["mask"].to(device)

        with autocast(enabled=use_amp):
            out       = model(images, return_features=False)
            loss_dict = criterion.supervised_loss(out, targets, model)

        if not is_valid_loss(loss_dict["total"]): continue

        val_loss += loss_dict["total"].item()
        pred      = keep_largest_component_tensor(out["pred"], threshold=0.5)
        metrics.update(pred, targets)
        n_steps += 1

        pbar.set_postfix({"val_loss": f"{loss_dict['total'].item():.4f}", "dice": f"{metrics.compute().get('dice', 0.0):.4f}"}, refresh=False)

    pbar.close()
    result = metrics.compute()
    result["val_loss"] = val_loss / max(n_steps, 1)
    return result


class ModelEMA:
    def __init__(self, model: nn.Module, decay: float = 0.99) -> None:
        self.decay   = decay
        self.shadow  = {name: param.data.clone() for name, param in model.named_parameters() if param.requires_grad}

    @torch.no_grad()
    def update(self, model: nn.Module) -> None:
        for name, param in model.named_parameters():
            if param.requires_grad and name in self.shadow:
                self.shadow[name] = self.decay * self.shadow[name] + (1.0 - self.decay) * param.data

    def apply(self, model: nn.Module):
        return _EMAContext(self, model)

class _EMAContext:
    def __init__(self, ema: ModelEMA, model: nn.Module) -> None:
        self.ema, self.model = ema, model
    def __enter__(self):
        self.backup = {name: param.data.clone() for name, param in self.model.named_parameters() if name in self.ema.shadow}
        for name, param in self.model.named_parameters():
            if name in self.ema.shadow: param.data.copy_(self.ema.shadow[name])
        return self.model
    def __exit__(self, *args):
        for name, param in self.model.named_parameters():
            if name in self.backup: param.data.copy_(self.backup[name])


def train(cfg: Config, debug: bool = False) -> None:
    if debug:
        cfg.train.num_epochs, cfg.train.batch_size, cfg.train.val_every, cfg.train.save_every, cfg.train.warmup_epochs = 3, 2, 1, 1, 1

    set_seed(cfg.train.seed)
    device = torch.device(cfg.train.device if torch.cuda.is_available() else "cpu")
    logger = setup_logging(cfg.train.log_dir)
    writer = SummaryWriter(cfg.train.log_dir)

    print_startup_table(cfg, device)


    logger.info("Building datasets...")
    train_loader, val_loader = build_labeled_loaders(cfg.data, cfg.train)
    logger.info(f"Train set: {len(train_loader.dataset)} samples | Val set: {len(val_loader.dataset)} samples")


    logger.info("Initializing CAFASegNet (DSBA-Bone backbone)...")

    model = CAFASegNet(
        in_channels      = cfg.model.in_channels,
        encoder_channels = cfg.model.encoder_channels,
        decoder_channels = cfg.model.decoder_channels,
        num_classes      = cfg.data.num_classes,
    ).to(device)

    logger.info(f"Model parameters: {sum(p.numel() for p in model.parameters())/1e6:.2f} M total")


    criterion = CAFASegNetLoss(cfg.train).to(device)
    backbone_lr_scale = getattr(cfg.train, "backbone_lr_scale", 1.0)

    if backbone_lr_scale != 1.0 and hasattr(model, "backbone"):

        pretrained_modules = [
            model.backbone.stem, model.backbone.layer1,
            model.backbone.layer2, model.backbone.layer3, model.backbone.layer4,
        ]
        pretrained_ids = set(id(p) for m in pretrained_modules for p in m.parameters())
        param_groups = [
            {"params": [p for p in model.parameters() if id(p) in pretrained_ids],
             "lr": cfg.train.lr * backbone_lr_scale, "name": "pretrained_semantic_stream"},
            {"params": [p for p in model.parameters() if id(p) not in pretrained_ids],
             "lr": cfg.train.lr, "name": "boundary_stream_and_decoder"},
        ]
        logger.info(f"Differential LR — semantic stream: {cfg.train.lr * backbone_lr_scale:.2e} (x{backbone_lr_scale})"
                    f" | boundary stream + decoder: {cfg.train.lr:.2e}")
    else:
        param_groups = list(model.parameters())
        logger.info(f"Uniform LR: lr={cfg.train.lr:.2e}")

    optimizer = torch.optim.AdamW(param_groups, lr=cfg.train.lr, weight_decay=cfg.train.weight_decay, betas=(0.9, 0.999))
    scheduler = build_scheduler(optimizer, cfg.train)

    ema = ModelEMA(model, decay=0.99)
    scaler = GradScaler(device=_AMP_DEVICE, enabled=cfg.train.amp) if _AMP_DEVICE else GradScaler(enabled=cfg.train.amp)
    excel_logger = ExcelLogger(cfg.train.log_dir)

    if cfg.train.resume:
        start_epoch, best_dice = load_checkpoint(model, optimizer, scheduler, scaler, cfg.train.resume, device, logger)
    else:
        start_epoch, best_dice = 0, 0.0

    global_step = 0
    best_path   = os.path.join(cfg.train.save_dir, "best_model.pth")

    print(f"{BOLD}{CYAN}{'━'*64}{RESET}\n{BOLD}{CYAN}   Start training  Epoch [{start_epoch+1} ~ {cfg.train.num_epochs}]{RESET}\n{BOLD}{CYAN}{'━'*64}{RESET}\n")

    epoch_pbar = tqdm(range(start_epoch, cfg.train.num_epochs), desc=f"{BOLD}Overall progress{RESET}", ncols=130, unit="ep",
                      bar_format=("  {desc}: {percentage:3.0f}%|{bar:36}| {n_fmt}/{total_fmt} ep  [{elapsed}<{remaining}]  {postfix}"))

    for epoch in epoch_pbar:
        t0 = time.time()
        tau = cosine_anneal_tau(epoch, cfg.train.tau_init, cfg.train.tau_min, cfg.train.tau_anneal_epochs)
        model.set_tau(tau)


        global_step, avg = train_one_epoch(model, train_loader, criterion, optimizer, scaler, device, cfg.train.amp, cfg.train.grad_clip, writer, global_step, epoch + 1, ema=ema)
        scheduler.step()
        elapsed = time.time() - t0

        print_train_table(epoch + 1, cfg.train.num_epochs, "SUP", avg, tau, get_lr(optimizer), elapsed)

        for k in ["total", "dice_loss", "tversky_loss", "bce_loss"]:
            writer.add_scalar(f"Epoch/Loss_{k}", avg[k], epoch)

        epoch_pbar.set_postfix({"loss": f"{avg['total']:.4f}", "dice▾": f"{avg['dice_loss']:.4f}"}, refresh=False)


        if (epoch + 1) % cfg.train.val_every == 0:
            with ema.apply(model):
                val_result = validate(model, val_loader, criterion, device, cfg.train.amp, epoch + 1, ema=ema)

            print_val_table(epoch + 1, val_result)
            logger.info(f"Epoch {epoch+1:3d} | TrainLoss={avg['total']:.4f} | VAL: Dice={val_result['dice']:.4f} HD95={val_result['hd95']:.1f}mm | LR={get_lr(optimizer):.2e}")

            excel_logger.append(epoch + 1, "SUP", avg, val_result, get_lr(optimizer), tau, elapsed)

            is_best = val_result["dice"] > best_dice
            if is_best:
                best_dice = val_result["dice"]
                print_best_banner(epoch + 1, best_dice)

            state = {
                "epoch":     epoch,
                "model":     model.state_dict(),
                "optimizer": optimizer.state_dict(),
                "scheduler": scheduler.state_dict(),
                "scaler":    scaler.state_dict(),
                "best_dice": best_dice,
            }


            latest_path = os.path.join(cfg.train.save_dir, "latest_model.pth")
            save_checkpoint(state, path=latest_path, is_best=is_best, best_path=best_path)

    epoch_pbar.close()
    excel_logger.close()
    writer.close()

    print(f"\n{BOLD}{CYAN}{'━'*64}{RESET}\n{BOLD}{GREEN}   Training completed.{RESET}\n{BOLD}{CYAN}{'━'*64}{RESET}")
    print(tabulate([["Best Dice", f"{BOLD}{GREEN}{best_dice:.4f}{RESET}"], ["Best checkpoint", best_path]], headers=["Item", "Value"], tablefmt="rounded_outline"))

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--debug",     action="store_true")
    parser.add_argument("--resume",    type=str, default=None)
    parser.add_argument("--data_root", type=str, default=None)
    args = parser.parse_args()

    if args.resume: cfg.train.resume = args.resume
    if args.data_root: cfg.data.dataset_root = args.data_root
    require_runtime_assets()
    train(cfg, debug=args.debug)
